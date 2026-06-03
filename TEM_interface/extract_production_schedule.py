"""
Targeted deep extraction of the Production Schedule sheet from the TEM Excel file.
Extracts every row with label, unit, formula, value and color meaning.
"""
import openpyxl

EXCEL_FILE = "/Users/rishidas/Documents/Internship_SRK/TEM_Mining/TEM_interface/Option 3_15 Mtpa_rev1_smg opt2 quality.xlsx"

print("Loading workbook with values...")
wb_v = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print("Loading workbook with formulas...")
wb_f = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

YEAR_COLS = {
    'E': -4, 'F': -3, 'G': -2, 'H': -1,  # Pre-production years
    'I': 1, 'J': 2, 'K': 3, 'L': 4, 'M': 5,
    'N': 6, 'O': 7, 'P': 8, 'Q': 9, 'R': 10,
    'S': 11, 'T': 12, 'U': 13, 'V': 14, 'W': 15,
    'X': 16, 'Y': 17, 'Z': 18, 'AA': 19
}

def get_fill_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb
    return None

HARD_INPUT_COLORS = {"FFFFFFCC"}  # Yellow = hard input in Production Schedule
RED_COLORS = {"FFFF0000", "FFC00000", "FF8B0000"}

ws_v = wb_v['Production Schedule']
ws_f = wb_f['Production Schedule']

print("\n" + "="*80)
print("PRODUCTION SCHEDULE SHEET — COMPLETE ROW-BY-ROW ANALYSIS")
print("="*80)
print(f"Dimensions: {ws_v.dimensions}")
print(f"Max rows: {ws_v.max_row}, Max cols: {ws_v.max_column}")

# Print every row that has content in col A, B, or C (row labels)
for row_num in range(1, ws_v.max_row + 1):
    row_has_data = False
    for col in range(1, ws_v.max_column + 1):
        cell = ws_v.cell(row_num, col)
        if cell.value is not None:
            row_has_data = True
            break
    if not row_has_data:
        continue

    # Get row label from col A (hard input rate), col B (description), col C (unit)
    a_val = ws_v.cell(row_num, 1).value
    b_val = ws_v.cell(row_num, 2).value
    c_val = ws_v.cell(row_num, 3).value
    d_val = ws_v.cell(row_num, 4).value
    a_formula = ws_f.cell(row_num, 1).value
    b_formula = ws_f.cell(row_num, 2).value

    print(f"\n--- ROW {row_num} ---")
    if a_val is not None:
        a_color = get_fill_color(ws_f.cell(row_num, 1))
        is_hard = a_color in HARD_INPUT_COLORS
        print(f"  [COL A] = {repr(a_val)!s:30} | COLOR={a_color} | HARD_INPUT={is_hard}")
        if a_formula and str(a_formula).startswith("="):
            print(f"           FORMULA: {a_formula}")
    if b_val is not None:
        print(f"  [LABEL/B] = {repr(b_val)}")
    if c_val is not None:
        print(f"  [UNIT/C] = {repr(c_val)}")
    if d_val is not None:
        d_formula = ws_f.cell(row_num, 4).value
        d_color = get_fill_color(ws_f.cell(row_num, 4))
        is_hard = d_color in HARD_INPUT_COLORS
        print(f"  [LOM TOTAL / D] = {repr(d_val)!s:30} | COLOR={d_color} | HARD_INPUT={is_hard}")
        if d_formula and str(d_formula).startswith("="):
            print(f"           FORMULA: {d_formula}")

    # Print year-by-year values with formulas for col I (Year 1) through AA (Year 19)
    # Also E,F,G,H for pre-production
    print(f"  Year-by-year data:")
    for col_letter, year_num in YEAR_COLS.items():
        if len(col_letter) == 1:
            col_idx = ord(col_letter) - ord('A') + 1
        else:  # AA
            col_idx = 27  # AA = column 27

        cell_v = ws_v.cell(row_num, col_idx)
        cell_f = ws_f.cell(row_num, col_idx)
        val = cell_v.value
        formula = cell_f.value
        color = get_fill_color(cell_f)
        is_hard = color in HARD_INPUT_COLORS

        if val is not None or (formula and str(formula).startswith("=")):
            formula_str = formula if (formula and str(formula).startswith("=")) else ""
            hard_marker = " <<HARD INPUT>>" if is_hard else ""
            print(f"    Year {year_num:+3d} ({col_letter:2s}): VALUE={repr(val)!s:20} FORMULA={formula_str!s:60}{hard_marker}")

# Also extract Assumptions_Dashboard cells that are referenced by Production Schedule
print("\n\n" + "="*80)
print("ASSUMPTIONS_DASHBOARD — CELLS REFERENCED FROM PRODUCTION SCHEDULE")
print("="*80)

# Key cells found in formulas
ws_dash_v = wb_v['Assumptions_Dashboard']
ws_dash_f = wb_f['Assumptions_Dashboard']

key_cells = [
    ("D5", "Effective Start Date for Model"),
    ("D6", "Production Start Date"),
    ("D7", "Production Starting Year"),
    ("D10", "Rated Capacity (Mtpa)"),
    ("D12", "Life of Mine (years)"),
    ("D13", "Block Area / ML Area (ha)"),
    ("D30", "Stripping Ratio (Overall)"),
    ("D31", "GCV Base"),
    ("D32", "Rated Capacity (Mtpa) — used in Rehandling calc"),
    ("D133", "Upfront Amount (INR Cr)"),
    ("D150", "MDO Top Soil rate (INR/bcm)"),
    ("D151", "MDO OB rate (INR/bcm)"),
    ("D152", "MDO Coal rate (INR/t)"),
    ("D153", "MDO Rehandling rate (INR/bcm)"),
    ("L4", "Mining Operation Type Switch"),
    ("L6", "Pre-Tax / Pre-Finance Switch"),
    ("L8", "Coal Price Type Switch"),
]

for cell_ref, desc in key_cells:
    cell_v = ws_dash_v[cell_ref]
    cell_f = ws_dash_f[cell_ref]
    color = get_fill_color(cell_f)
    formula = cell_f.value
    is_hard = color in HARD_INPUT_COLORS | RED_COLORS | HARD_INPUT_COLORS
    print(f"  [{cell_ref}] {desc}")
    print(f"       VALUE = {repr(cell_v.value)}")
    if formula and str(formula).startswith("="):
        print(f"       FORMULA = {formula}")
    print(f"       COLOR = {color}")
    print()

print("\n\nDone.")
