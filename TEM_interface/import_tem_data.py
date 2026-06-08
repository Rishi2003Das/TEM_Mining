import json
import os
import openpyxl
import pymongo

# MongoDB connection configuration
CONNECTION_STRING = "mongodb+srv://rishikakalidas:KNris$0068@tem.khdmanp.mongodb.net/?appName=tem"
DATABASE_NAME = "tem"
EXCEL_FILE = "Option 3_15 Mtpa_rev1_smg opt2 quality.xlsx"

# Mapping of JSON files to collection names
JSON_FILES_MAP = {
    "salary_wages.json": "salary_wages",
    "working_regime.json": "working_regime",
    "basic_consideration.json": "basic_consideration",
    "density_swell_factor.json": "density_swell_factor",
    "explosives.json": "explosives",
    "unit_rates_opcosts.json": "unit_rates_opcosts",
    "maintainance_cost.json": "maintainance_cost",
    "operational_para.json": "operational_para",
    "govt_fees_charges.json": "govt_fees_charges",
    "payment_assumption.json": "payment_assumption",
    "mdo_assumption.json": "mdo_assumption",
    "safety_slope_stability.json": "safety_slope_stability"
}

# Excel sheet mappings to import
# Columns for year headers in the sheets (E to AA are columns 5 to 27)
# Year header mapping: col 5 is -4, col 6 is -3, ..., col 27 is 19
YEAR_HEADERS = [-4, -3, -2, -1, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
COL_START = 5 # Col E
COL_END = 27  # Col AA

def parse_schedule_row(sheet, row_idx, description_col=2, unit_col=3, total_col=4, col_start=5, col_end=27):
    """
    Parse a single row from a year-by-year schedule sheet.
    """
    description = sheet.cell(row_idx, description_col).value
    if not description:
        return None
    
    unit = sheet.cell(row_idx, unit_col).value
    total_val = sheet.cell(row_idx, total_col).value
    
    yearly_values = {}
    for col_idx in range(col_start, col_end + 1):
        year_name = str(YEAR_HEADERS[col_idx - col_start])
        val = sheet.cell(row_idx, col_idx).value
        # Convert Excel None or formula strings/empty to 0.0 or float
        if val is None or val == "":
            yearly_values[year_name] = 0.0
        elif isinstance(val, (int, float)):
            yearly_values[year_name] = float(val)
        else:
            try:
                yearly_values[year_name] = float(val)
            except ValueError:
                yearly_values[year_name] = val # Keep as string if it is not numeric
                
    return {
        "item": str(description).strip(),
        "unit": str(unit).strip() if unit else "",
        "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
        "yearly_values": yearly_values
    }

def import_tem_data():
    print("Connecting to MongoDB...")
    try:
        client = pymongo.MongoClient(CONNECTION_STRING)
        db = client[DATABASE_NAME]
        client.admin.command('ping')
        print("Successfully connected to MongoDB.")
    except Exception as e:
        print(f"Error connecting to MongoDB: {e}")
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    hard_input_dir = os.path.join(script_dir, "Hard_Input")

    # 1. Import JSON files
    for filename, collection_name in JSON_FILES_MAP.items():
        file_path = os.path.join(hard_input_dir, filename)
        if not os.path.exists(file_path):
            print(f"Error: File {filename} not found in {hard_input_dir}. Skipping.")
            continue

        print(f"Importing {filename} into collection '{collection_name}'...")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not isinstance(data, list):
                data = [data]

            collection = db[collection_name]
            collection.delete_many({})
            if data:
                collection.insert_many(data)
                print(f"Successfully inserted {len(data)} documents into '{collection_name}'.")
        except Exception as e:
            print(f"Failed to import {filename}: {e}")

    # 2. Parse and import Excel schedules
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel workbook {EXCEL_FILE} not found. Cannot parse schedules.")
        return

    print(f"\nLoading Excel workbook: {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # A. Production Schedule Sheet
    print("Parsing 'Production Schedule'...")
    prod_sheet = wb["Production Schedule"]
    prod_docs = []
    # Rows to parse: 4 (Total available hours), 5 (Production coal), 8 (Waste), 9 (Top Soil), 10 (OB), 11 (Partings), 12 (Rehandling CHP), 13 (Waste Rehandle), 14 (GCV), 15 (RD), 16 (Raw Ash), 17 (Moisture), 23 (ROM), 24 (Waste), 25 (In-Pit), 26 (Ex-Pit), 27 (Rehandling), 30 (Bench height coal), 31 (Bench width coal), 32 (Bench height OB/IB), 33 (Bench width OB/IB), 35 (Rehandling cost)
    prod_rows = [4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 23, 24, 25, 26, 27, 30, 31, 32, 33, 35]
    for r in prod_rows:
        doc = parse_schedule_row(prod_sheet, r)
        if doc:
            prod_docs.append(doc)
    
    # B. Pre-Operative Sheet
    print("Parsing 'Pre-Operative'...")
    pre_sheet = wb["Pre-Operative"]
    pre_docs = []
    # Parse rows 5 to 38. Pre-production years are -4 to -1 (cols 6 to 9 in Pre-Operative)
    # Let's adjust parser for Pre-Operative column headers (cols 6 to 9 correspond to years -4 to -1)
    for r in range(5, 41):
        description = pre_sheet.cell(r, 2).value
        if not description:
            continue
        unit = pre_sheet.cell(r, 3).value
        total_val = pre_sheet.cell(r, 5).value
        yearly_values = {}
        # Pre-operative sheet has -4, -3, -2, -1 in columns 6, 7, 8, 9
        pre_years = [-4, -3, -2, -1]
        for idx, yr in enumerate(pre_years):
            val = pre_sheet.cell(r, 6 + idx).value
            yearly_values[str(yr)] = float(val) if isinstance(val, (int, float)) else (0.0 if val is None else val)
        # Pad other years with 0.0
        for yr in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
            yearly_values[str(yr)] = 0.0
            
        pre_docs.append({
            "item": str(description).strip(),
            "unit": str(unit).strip() if unit else "",
            "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
            "yearly_values": yearly_values
        })

    # C. Land Sheet
    print("Parsing 'Land'...")
    land_sheet = wb["Land"]
    land_docs = []
    # Row 5 (Land Costs in INR Crore), Row 6 (multiplication factor series)
    for r in [5, 6]:
        doc = parse_schedule_row(land_sheet, r, description_col=2, unit_col=2, total_col=3, col_start=4, col_end=26)
        if doc:
            if r == 6:
                doc["item"] = "Land Cost Factor"
            land_docs.append(doc)

    # D. R&R Sheet
    print("Parsing 'R&R'...")
    rr_sheet = wb["R&R"]
    rr_docs = []
    # Row 6 (R&R cost in INR Crore), Row 7 (Opex R&R cost)
    for r in [6, 7]:
        doc = parse_schedule_row(rr_sheet, r, description_col=2, unit_col=3, total_col=4, col_start=4, col_end=26)
        if doc:
            rr_docs.append(doc)

    # E. Coal Price Sheet
    print("Parsing 'Coal Price'...")
    cp_sheet = wb["Coal Price"]
    cp_docs = []
    # Row 5 (GCV), Row 7 (NCI Price), Row 8 (Market Price/Commercial)
    for r in [5, 7, 8]:
        doc = parse_schedule_row(cp_sheet, r, description_col=2, unit_col=3, total_col=4)
        if doc:
            cp_docs.append(doc)

    # F. Capex Breakups Sheet
    print("Parsing 'Capex Breakups'...")
    cb_sheet = wb["Capex Breakups"]
    cb_docs = []
    # Rows 4 to 11 (CHP, Railway Siding, Civil, Fire Fighting, Workshop, Electrical, Mine Dewatering, Digitalisation)
    for r in range(4, 12):
        doc = parse_schedule_row(cb_sheet, r, description_col=2, unit_col=3, total_col=3)
        if doc:
            cb_docs.append(doc)

    # G. Fleet Sheet
    print("Parsing 'Fleet'...")
    fleet_sheet = wb["Fleet"]
    fleet_docs = []
    # Row 366 (Total Initial Capital Requirement - Equipment)
    # Row 413 (Total Replacement Capital Requirement - Equipment)
    # Note: Fleet sheet year columns starting from -4 to 19 correspond to columns 7 (G) to 29 (AC)
    for r in [366, 413]:
        description = fleet_sheet.cell(r, 3).value
        unit = fleet_sheet.cell(r, 4).value
        total_val = fleet_sheet.cell(r, 6).value
        yearly_values = {}
        for col_idx in range(7, 30):
            year_name = str(YEAR_HEADERS[col_idx - 7])
            val = fleet_sheet.cell(r, col_idx).value
            yearly_values[year_name] = float(val) if isinstance(val, (int, float)) else 0.0

        fleet_docs.append({
            "item": str(description).strip(),
            "unit": str(unit).strip() if unit else "",
            "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
            "yearly_values": yearly_values
        })

    # Wage-Owner Sheet
    print("Parsing 'Wage-Owner'...")
    wage_sheet = wb["Wage-Owner"]
    # Row 280 (Total Wage Cost in INR Cr)
    # Note: Wage-Owner year columns starting from -4 to 19 correspond to columns 10 (J) to 32 (AF)
    description = "Total Wages"
    unit = "INR Cr"
    total_val = wage_sheet.cell(280, 6).value # Total column or average
    yearly_values = {}
    for col_idx in range(10, 33):
        year_name = str(YEAR_HEADERS[col_idx - 10])
        val = wage_sheet.cell(280, col_idx).value
        yearly_values[year_name] = float(val) if isinstance(val, (int, float)) else 0.0

    wage_docs = [{
        "item": description,
        "unit": unit,
        "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
        "yearly_values": yearly_values
    }]

    # H. Government Sheet (Actual computed schedules for regression testing)
    print("Parsing 'Government'...")
    gov_sheet = wb["Government"]
    gov_docs = []
    # Rows 4 to 36 (Coal production, NCI price, commercial price, revenue sharing, royalties, etc.)
    for r in range(4, 37):
        doc = parse_schedule_row(gov_sheet, r, description_col=2, unit_col=3, total_col=4)
        if doc:
            gov_docs.append(doc)

    # I. Owner OPEX (Actual computed opex for regression testing)
    print("Parsing 'Owner OPEX'...")
    owner_opex_sheet = wb["Owner OPEX"]
    owner_opex_docs = []
    for r in range(56, 93):
        doc = parse_schedule_row(owner_opex_sheet, r, description_col=2, unit_col=3, total_col=4)
        if doc:
            owner_opex_docs.append(doc)

    # J. Project OPEX (Actual computed project opex for regression testing)
    print("Parsing 'Project OPEX'...")
    project_opex_sheet = wb["Project OPEX"]
    project_opex_docs = []
    for r in range(59, 101):
        doc = parse_schedule_row(project_opex_sheet, r, description_col=2, unit_col=3, total_col=4)
        if doc:
            project_opex_docs.append(doc)

    # Write schedules to database
    db_mappings = {
        "production_schedule": prod_docs,
        "pre_operative_schedule": pre_docs,
        "land_schedule": land_docs,
        "rr_schedule": rr_docs,
        "coal_price_schedule": cp_docs,
        "capex_breakups_schedule": cb_docs,
        "fleet_replacement_schedule": fleet_docs,
        "wages_schedule": wage_docs,
        "government_schedule": gov_docs,
        "owner_opex_schedule": owner_opex_docs,
        "project_opex_schedule": project_opex_docs
    }

    for col_name, docs in db_mappings.items():
        print(f"Importing {len(docs)} documents into collection '{col_name}'...")
        collection = db[col_name]
        collection.delete_many({})
        if docs:
            collection.insert_many(docs)
            print(f"Successfully inserted into '{col_name}'.")

    print("\nAll spreadsheet and JSON data imported to MongoDB successfully.")

if __name__ == "__main__":
    import_tem_data()
