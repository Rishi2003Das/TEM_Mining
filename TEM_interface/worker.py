import sys
import os
import json
import openpyxl
import pymongo
import datetime

# Configuration
CONNECTION_STRING = os.environ.get("MONGODB_URI", "mongodb+srv://rishikakalidas:KNris$0068@tem.khdmanp.mongodb.net/?appName=tem")
DATABASE_NAME = os.environ.get("DATABASE_NAME", "tem")

# Hard input sheets mapping
HARD_INPUT_MAPPING = {
    "basic_consideration": {"json_file": "basic_consideration.json", "val_field": "Value"},
    "working_regime": {"json_file": "working_regime.json", "val_field": "Values"},
    "density_swell_factor": {"json_file": "density_swell_factor.json", "val_field": "Value"},
    "salary_wages": {"json_file": "salary_wages.json", "val_field": "Annual CTC"},
    "explosives": {"json_file": "explosives.json", "val_field": "Value"},
    "unit_rates_opcosts": {"json_file": "unit_rates_opcosts.json", "val_field": "Base Rate"},
    "maintainance_cost": {"json_file": "maintainance_cost.json", "val_field": "Base Rate"},
    "operational_para": {"json_file": "operational_para.json", "val_field": "Values"},
    "govt_fees_charges": {"json_file": "govt_fees_charges.json", "val_field": "Base Rate"},
    "payment_assumption": {"json_file": "payment_assumption.json", "val_field": "Base Rate"},
    "mdo_assumption": {"json_file": "mdo_assumption.json", "val_field": "Value"},
    "safety_slope_stability": {"json_file": "safety_slope_stability.json", "val_field": "Value"}
}

# Section headers inside Assumptions_Dashboard
SECTION_HEADERS = {
    "basic_consideration": "Basic Considerations",
    "working_regime": "Working Regime",
    "density_swell_factor": "Density & Swell Factor",
    "salary_wages": "Salary & Wages",
    "explosives": "Explosives",
    "unit_rates_opcosts": "Unit Rates and Operating Costs",
    "maintainance_cost": "Maintenance Rate",
    "operational_para": "Operational Parameters",
    "govt_fees_charges": "Government Fees and charges",
    "payment_assumption": "Payment related assumption",
    "mdo_assumption": "MDO Assumption",
    "safety_slope_stability": "Safety - slope stability monitoring"
}

YEAR_HEADERS = [-4, -3, -2, -1, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]

def find_row_by_label(sheet, label, col_idx=2, start_row=1, max_rows=1000, fuzzy=True):
    target = label.strip().lower()
    for r in range(start_row, max_rows + 1):
        val = sheet.cell(r, col_idx).value
        if val is not None:
            val_str = str(val).strip().lower()
            if fuzzy:
                if target in val_str or val_str in target:
                    return r
            else:
                if target == val_str:
                    return r
    return None

def get_param_value(wb, sheet_name, search_label, target_col=4, search_col=2, fuzzy=True):
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    row_idx = find_row_by_label(sheet, search_label, col_idx=search_col, fuzzy=fuzzy)
    if row_idx is not None:
        return sheet.cell(row_idx, target_col).value
    return None

def get_description_field_name(item):
    # Description field is any key that is not metadata or value fields
    ignored_keys = {"id", "key", "Unit", "Value", "Values", "Annual CTC", "Base Rate", "Basis", "Representative Price", "Operational Parameters"}
    for k in item.keys():
        if k not in ignored_keys:
            return k
    if "Operational Parameters" in item:
        return "Operational Parameters"
    return None

def parse_schedule_row(sheet, row_idx, description_col=2, unit_col=3, total_col=4, col_start=5, col_end=27, item_override=None):
    description = item_override or sheet.cell(row_idx, description_col).value
    if not description:
        return None
    
    unit = sheet.cell(row_idx, unit_col).value
    total_val = sheet.cell(row_idx, total_col).value
    
    yearly_values = {}
    for col_idx in range(col_start, col_end + 1):
        year_name = str(YEAR_HEADERS[col_idx - col_start])
        val = sheet.cell(row_idx, col_idx).value
        if val is None or val == "":
            yearly_values[year_name] = 0.0
        elif isinstance(val, (int, float)):
            yearly_values[year_name] = float(val)
        else:
            try:
                yearly_values[year_name] = float(val)
            except ValueError:
                yearly_values[year_name] = val
                
    return {
        "item": str(description).strip(),
        "unit": str(unit).strip() if unit else "",
        "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
        "yearly_values": yearly_values
    }

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 worker.py <path_to_excel_file>")
        sys.exit(1)
        
    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
        
    print(f"Connecting to MongoDB...")
    client = pymongo.MongoClient(CONNECTION_STRING)
    db = client[DATABASE_NAME]
    
    # Verify connection
    client.admin.command('ping')
    print("MongoDB connected successfully.")
    
    # Update project metadata from environment variables if present
    project_id = os.environ.get("PROJECT_ID")
    project_manager = os.environ.get("PROJECT_MANAGER")
    client_company = os.environ.get("CLIENT_COMPANY")
    project_description = os.environ.get("PROJECT_DESCRIPTION")
    
    if project_id:
        print(f"Saving project metadata for ID: {project_id}...")
        db["project_metadata"].delete_many({})
        db["project_metadata"].insert_one({
            "projectId": project_id,
            "projectManager": project_manager or "",
            "clientCompany": client_company or "",
            "projectDescription": project_description or ""
        })
        print("Project metadata successfully saved.")
        
    print(f"Loading Excel workbook: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    hard_input_dir = os.path.join(script_dir, "Hard_Input")
    
    # 1. Parse and extract assumptions from 'Assumptions_Dashboard'
    print("\nParsing 'Assumptions_Dashboard' for hard inputs dynamically...")
    dash_sheet = wb["Assumptions_Dashboard"]
    
    for collection_name, config in HARD_INPUT_MAPPING.items():
        json_path = os.path.join(hard_input_dir, config["json_file"])
        if not os.path.exists(json_path):
            print(f"Warning: JSON template {config['json_file']} not found. Skipping.")
            continue
            
        with open(json_path, 'r', encoding='utf-8') as f:
            template_items = json.load(f)
            
        section_header = SECTION_HEADERS.get(collection_name)
        sec_start_row = 1
        if section_header:
            row_match = find_row_by_label(dash_sheet, section_header, col_idx=2, max_rows=300, fuzzy=True)
            if row_match:
                sec_start_row = row_match
                
        updated_items = []
        for item in template_items:
            desc_field = get_description_field_name(item)
            if not desc_field:
                print(f"Warning: No description field found for item key: {item.get('key')}")
                updated_items.append(item)
                continue
                
            desc_val = str(item[desc_field]).strip()
            
            # Find row index inside section
            row_idx = None
            for r in range(sec_start_row, min(sec_start_row + 60, 500)):
                val_b = dash_sheet.cell(r, 2).value
                if val_b and desc_val.lower().strip() in str(val_b).lower().strip():
                    row_idx = r
                    break
                    
            if row_idx is None:
                # Fallback to scan entire sheet
                row_idx = find_row_by_label(dash_sheet, desc_val, col_idx=2, max_rows=300, fuzzy=True)
                
            if row_idx is not None:
                val = dash_sheet.cell(row_idx, 4).value
                
                # Format cell value
                if isinstance(val, (int, float)):
                    val = float(val)
                elif val is None:
                    val = 0.0
                elif hasattr(val, "strftime"):  # datetime
                    val = val.strftime("%d-%b-%y")
                else:
                    val = str(val).strip()
                    try:
                        val = float(val.replace(",", ""))
                    except ValueError:
                        pass
                
                item[config["val_field"]] = val
            else:
                print(f"Warning: Label '{desc_val}' not found in Assumptions_Dashboard. Keeping default template value.")
                
            updated_items.append(item)
            
        print(f"Upserting {len(updated_items)} items into collection '{collection_name}'...")
        db[collection_name].delete_many({})
        if updated_items:
            db[collection_name].insert_many(updated_items)
            
    # 2. Parse and extract 'production_schedule_params' dynamically
    print("\nParsing production schedule parameters dynamically...")
    params_json_path = os.path.join(hard_input_dir, "production_schedule_params.json")
    if os.path.exists(params_json_path):
        with open(params_json_path, 'r', encoding='utf-8') as f:
            params_template = json.load(f)
            
        # Label search mapping: key -> (sheet_name, search_label, target_col, search_col, fuzzy)
        param_search_mapping = {
            "partings_percent": ("Production Schedule", "Partings", 1, 2, False),
            "chp_rehandling_rate": ("Production Schedule", "Rehandling at CHP", 1, 2, False),
            "chp_rehandling_capacity": ("Assumptions_Dashboard", "Average density of Coal", 4, 2, True),
            "blasted_coal_fraction": ("Assumptions_Dashboard", "Blasting Requirement in coal through out life", 4, 2, True),
            "sm_threshold_year": ("Assumptions_Dashboard", "Blasting Requirement 100% for how many years", 4, 2, True),
            "available_hours_per_year": ("Assumptions_Dashboard", "Total available Hours", 4, 2, True),
            "bench_height_coal": ("Production Schedule", "Bench Height - Coal", 4, 2, True),
            "bench_width_coal": ("Production Schedule", "Bench Width - Coal", 4, 2, True),
            "bench_height_ob_ib": ("Production Schedule", "Bench Height - OB/IB", 4, 2, True),
            "bench_width_ob_ib": ("Production Schedule", "Bench Width - OB/IB", 4, 2, True)
        }
        
        for item in params_template:
            key = item.get("key")
            if key in param_search_mapping:
                sheet_name, search_label, target_col, search_col, fuzzy = param_search_mapping[key]
                val = get_param_value(wb, sheet_name, search_label, target_col, search_col, fuzzy)
                if val is not None:
                    if isinstance(val, (int, float)):
                        val = float(val)
                    else:
                        try:
                            # Handle percentage string
                            val = float(str(val).replace("%", "").strip())
                            if "%" in str(get_param_value(wb, sheet_name, search_label, target_col, search_col, fuzzy) or ""):
                                val = val / 100.0
                        except ValueError:
                            pass
                    item["Value"] = val
                else:
                    print(f"Warning: Dynamic value for param '{key}' not found.")
                    
        print(f"Upserting production_schedule_params...")
        db["production_schedule_params"].delete_many({})
        db["production_schedule_params"].insert_many(params_template)

    # 3. Parse and extract schedules dynamically (year-by-year)
    # A. Production Schedule
    print("\nParsing 'Production Schedule' dynamically...")
    prod_sched_sheet = wb["Production Schedule"]
    PROD_SCHED_ROW_MAPPING = {
        "total available hours excavation": "Total available Hours excavation",
        "production coal/ore": "Production Coal/Ore",
        "blasted coal/ore": "Blasted Coal/Ore",
        "surface miner coal/ore": "Surface Miner Coal/Ore",
        "waste (topsoil + overburden + interburden)": "Waste (Topsoil + Overburden + Interburden)",
        "waste (topsoil + ob + partings)": "Waste (Topsoil + Overburden + Interburden)",
        "top soil": "Topsoil",
        "topsoil": "Topsoil",
        "ob": "OB",
        "partings": "Partings",
        "rehandling at chp": "Rehandling at CHP",
        "waste - rehandle": "Waste - Rehandle",
        "gcv (adb)": "GCV (ADB)",
        "relaive density (rd)": "Relaive Density (RD)",
        "relative density": "Relaive Density (RD)",
        "raw ash": "Raw ash",
        "moisture": "Moisture",
        "rom": "ROM",
        "waste": "Waste",
        "in-pit": "In-Pit",
        "ex-pit": "Ex-Pit",
        "rehandling": "Rehandling",
        "bench height - coal": "Bench Height - Coal",
        "bench width - coal": "Bench Width - Coal",
        "bench height - ob/ib": "Bench Height - OB/IB",
        "bench width - ob/ib": "Bench Width - OB/IB",
        "rehandling cost": "Rehandling cost"
    }
    prod_docs = []
    for r in range(1, 150):
        val_b = prod_sched_sheet.cell(r, 2).value
        if val_b:
            key_b = str(val_b).strip().lower()
            if key_b in PROD_SCHED_ROW_MAPPING:
                std_name = PROD_SCHED_ROW_MAPPING[key_b]
                # Avoid duplicate rows
                if any(d["item"] == std_name for d in prod_docs):
                    continue
                doc = parse_schedule_row(prod_sched_sheet, r)
                if doc:
                    doc["item"] = std_name
                    prod_docs.append(doc)
                    
    # B. Pre-Operative
    print("Parsing 'Pre-Operative' dynamically...")
    pre_sheet = wb["Pre-Operative"]
    pre_docs = []
    header_row = find_row_by_label(pre_sheet, "Description", col_idx=2)
    if header_row:
        for r in range(header_row + 1, 200):
            val_b = pre_sheet.cell(r, 2).value
            if not val_b:
                continue
            val_b_str = str(val_b).strip().lower()
            
            description = val_b
            unit = pre_sheet.cell(r, 3).value
            total_val = pre_sheet.cell(r, 5).value
            
            # Map pre-operative columns F to I for years -4 to -1
            yearly_values = {}
            pre_years = [-4, -3, -2, -1]
            for idx, yr in enumerate(pre_years):
                val = pre_sheet.cell(r, 6 + idx).value
                yearly_values[str(yr)] = float(val) if isinstance(val, (int, float)) else (0.0 if val is None else val)
            for yr in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                yearly_values[str(yr)] = 0.0
                
            pre_docs.append({
                "item": str(description).strip(),
                "unit": str(unit).strip() if unit else "",
                "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
                "yearly_values": yearly_values
            })
            if val_b_str == "total":
                break
                
    # C. Land
    print("Parsing 'Land' dynamically...")
    land_sheet = wb["Land"]
    land_row = find_row_by_label(land_sheet, "Land Costs, (INR Crore)", col_idx=2, fuzzy=False)
    land_docs = []
    if land_row:
        doc1 = parse_schedule_row(land_sheet, land_row, description_col=2, unit_col=2, total_col=3, col_start=4, col_end=26)
        doc2 = parse_schedule_row(land_sheet, land_row + 1, description_col=2, unit_col=2, total_col=3, col_start=4, col_end=26, item_override="Land Cost Factor")
        if doc1:
            land_docs.append(doc1)
        if doc2:
            land_docs.append(doc2)
            
    # D. R&R
    print("Parsing 'R&R' dynamically...")
    rr_sheet = wb["R&R"]
    rr_row = find_row_by_label(rr_sheet, "R&R", col_idx=2, fuzzy=False)
    rr_docs = []
    if rr_row:
        doc1 = parse_schedule_row(rr_sheet, rr_row, description_col=2, unit_col=3, total_col=4, col_start=4, col_end=26)
        doc2 = parse_schedule_row(rr_sheet, rr_row + 1, description_col=2, unit_col=3, total_col=4, col_start=4, col_end=26)
        if doc1:
            rr_docs.append(doc1)
        if doc2:
            rr_docs.append(doc2)
            
    # E. Coal Price
    print("Parsing 'Coal Price' dynamically...")
    cp_sheet = wb["Coal Price"]
    cp_labels = ["GCV as per Mine Design", "NCI Price as per Mine Design Grade", "Market Price/Commercial Price of Coal"]
    cp_docs = []
    for label in cp_labels:
        r = find_row_by_label(cp_sheet, label, col_idx=2)
        if r:
            doc = parse_schedule_row(cp_sheet, r, description_col=2, unit_col=3, total_col=4)
            if doc:
                cp_docs.append(doc)
                
    # F. Capex Breakups
    print("Parsing 'Capex Breakups' dynamically...")
    cb_sheet = wb["Capex Breakups"]
    cb_header = find_row_by_label(cb_sheet, "Item", col_idx=2)
    cb_docs = []
    if cb_header:
        for r in range(cb_header + 1, 100):
            val_b = cb_sheet.cell(r, 2).value
            if not val_b:
                continue
            if str(val_b).strip().lower() == "total":
                break
            doc = parse_schedule_row(cb_sheet, r, description_col=2, unit_col=3, total_col=3)
            if doc:
                cb_docs.append(doc)
                
    # G. Fleet
    print("Parsing 'Fleet' dynamically...")
    fleet_sheet = wb["Fleet"]
    fleet_docs = []
    for label in ["Total Initial Capital Requirement", "Total Replacement Capital Requirement"]:
        r = find_row_by_label(fleet_sheet, label, col_idx=3)
        if r:
            description = fleet_sheet.cell(r, 3).value
            unit = fleet_sheet.cell(r, 4).value
            total_val = fleet_sheet.cell(r, 6).value
            yearly_values = {}
            for col_idx in range(7, 30):
                year_name = str(YEAR_HEADERS[col_idx - 7])
                val = fleet_sheet.cell(r, col_idx).value
                yearly_values[year_name] = float(val) if isinstance(val, (int, float)) else 0.0

            fleet_docs.append({
                "item": str(description).strip(),
                "unit": str(unit).strip() if unit else "",
                "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
                "yearly_values": yearly_values
            })
            
    # Wage-Owner
    print("Parsing 'Wage-Owner' dynamically...")
    wage_sheet = wb["Wage-Owner"]
    r_wage = find_row_by_label(wage_sheet, "Total", col_idx=3)
    wage_docs = []
    if r_wage:
        total_val = wage_sheet.cell(r_wage, 6).value
        yearly_values = {}
        for col_idx in range(10, 33):
            year_name = str(YEAR_HEADERS[col_idx - 10])
            val = wage_sheet.cell(r_wage, col_idx).value
            yearly_values[year_name] = float(val) if isinstance(val, (int, float)) else 0.0

        wage_docs.append({
            "item": "Total Wages",
            "unit": "INR Cr",
            "lom_total_or_average": float(total_val) if isinstance(total_val, (int, float)) else total_val,
            "yearly_values": yearly_values
        })
        
    # H. Government
    print("Parsing 'Government' dynamically...")
    gov_sheet = wb["Government"]
    gov_docs = []
    header_row_gov = find_row_by_label(gov_sheet, "Description", col_idx=2)
    if header_row_gov:
        empty_cnt = 0
        for r in range(header_row_gov + 1, 200):
            val_b = gov_sheet.cell(r, 2).value
            if not val_b:
                empty_cnt += 1
                if empty_cnt > 3:
                    break
                continue
            empty_cnt = 0
            doc = parse_schedule_row(gov_sheet, r, description_col=2, unit_col=3, total_col=4)
            if doc:
                gov_docs.append(doc)
                
    # I. Owner OPEX & J. Project OPEX
    owner_opex_docs = []
    project_opex_docs = []
    
    for op_name, op_sheet_name, op_list in [("Owner OPEX", "Owner OPEX", owner_opex_docs), ("Project OPEX", "Project OPEX", project_opex_docs)]:
        print(f"Parsing '{op_name}' dynamically...")
        op_sheet = wb[op_sheet_name]
        header_row_op = find_row_by_label(op_sheet, "Description", col_idx=2)
        if header_row_op:
            for r in range(header_row_op + 1, 250):
                val_b = op_sheet.cell(r, 2).value
                if not val_b:
                    continue
                # Skip header groupings
                val_c = op_sheet.cell(r, 3).value
                val_d = op_sheet.cell(r, 4).value
                if val_c is None and val_d is None:
                    continue
                doc = parse_schedule_row(op_sheet, r, description_col=2, unit_col=3, total_col=4)
                if doc:
                    op_list.append(doc)
                if str(val_b).strip().lower() == "grand total":
                    break

    # Save to DB
    db_mappings = {
        "production_schedule": prod_docs,
        "pre_operative_schedule": pre_docs,
        "land_schedule": land_docs,
        "rr_schedule": rr_docs,
        "coal_price_schedule": cp_docs,
        "capex_breakups_schedule": cb_docs,
        "fleet_replacement_schedule": fleet_docs,
        "wages_schedule": wage_docs,
        "government_schedule": gov_docs,
        "owner_opex_schedule": owner_opex_docs,
        "project_opex_schedule": project_opex_docs
    }

    for col_name, docs in db_mappings.items():
        print(f"Importing {len(docs)} documents into collection '{col_name}'...")
        collection = db[col_name]
        collection.delete_many({})
        if docs:
            collection.insert_many(docs)

    print("\nExcel data successfully imported to MongoDB dynamically.")
    
    # 4. Trigger calculate_tem.py
    print("Triggering recalculation...")
    import subprocess
    calc_script = os.path.join(script_dir, "calculate_tem.py")
    res = subprocess.run(["python3", calc_script], capture_output=True, text=True)
    if res.returncode == 0:
        print("Recalculation finished successfully.")
        print(res.stdout)
    else:
        print("Recalculation failed:")
        print(res.stderr)
        sys.exit(res.returncode)

if __name__ == "__main__":
    main()
