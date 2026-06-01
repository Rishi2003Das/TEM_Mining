import openpyxl

wb = openpyxl.load_workbook("Option 3_15 Mtpa_rev1_smg opt2 quality.xlsx", data_only=False)
sheet = wb["Owner CAPEX"]

print("Owner CAPEX Row 16 (CHP) Formulas:")
for col in range(5, 15):
    col_letter = openpyxl.utils.get_column_letter(col)
    year = wb["Owner CAPEX"].cell(10, col).value
    val = sheet.cell(16, col).value
    print(f"  Col {col_letter} (Year {year}): {val}")
