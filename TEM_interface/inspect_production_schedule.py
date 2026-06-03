"""
Deep inspection script for Production Schedule sheet in the TEM Excel file.
Extracts all cell values, formulas, inter-sheet links, and hard-input information.
"""
import openpyxl
from openpyxl.styles import PatternFill
import json, sys

EXCEL_FILE = "/Users/rishidas/Documents/Internship_SRK/TEM_Mining/TEM_interface/Option 3_15 Mtpa_rev1_smg opt2 quality.xlsx"

print("Loading workbook (read-only for speed)...")
wb_values = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print("Loading workbook (with formulas)...")
wb_formulas = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

print("\n=== SHEET NAMES ===")
for name in wb_values.sheetnames:
    print(f"  - {name}")

# ──────────────────────────────────────────────────────────────────────────────
# Helper: get fill colour of a cell
# ──────────────────────────────────────────────────────────────────────────────
def get_fill_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb  # e.g. "FFFF0000" for red
    return None

COLOR_MEANING = {
    "FF4472C4": "Blue (cross-sheet link)",
    "FF0070C0": "Blue (cross-sheet link)",
    "FFFF0000": "Red (hard input)",
    "FFFF0000": "Red (hard input)",
    "FFFFCCCC": "Pink (external file link)",
    "FF00B050": "Green (named value/array calc)",
    "FFBFBFBF": "Grey (not required)",
    "FFFFFF00": "Yellow (title/revision)",
    "FFED7D31": "Orange",
    "FF70AD47": "Green",
    "FFFFC000": "Amber/Yellow",
    # lighter shades
    "FFD6E4BC": "Light Green",
    "FFDAE3F3": "Light Blue",
    "FFFCE4D6": "Light Orange",
    "FFE2EFDA": "Light Green 2",
    "FFFFF2CC": "Light Yellow",
    "FFDD8888": "Light Red/Pink",
    "FFC00000": "Dark Red (hard input)",
}

def color_label(hex_color):
    if hex_color is None:
        return "No fill"
    return COLOR_MEANING.get(hex_color, f"Unknown ({hex_color})")

# ──────────────────────────────────────────────────────────────────────────────
# 1. GUIDE SHEET — dump all non-empty cells
# ──────────────────────────────────────────────────────────────────────────────
print("\n\n=== GUIDE SHEET — ALL NON-EMPTY CELLS ===\n")
guide_names = [n for n in wb_values.sheetnames if "guide" in n.lower() or "legend" in n.lower() or "colour" in n.lower()]
print(f"  Guide-like sheets: {guide_names}")

for gname in guide_names:
    ws = wb_values[gname]
    wsf = wb_formulas[gname]
    print(f"\n--- Sheet: {gname} ---")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                formula_cell = wsf[cell.coordinate]
                formula = formula_cell.value if str(formula_cell.value).startswith("=") else ""
                color = get_fill_color(wsf[cell.coordinate])
                print(f"  {cell.coordinate}: VALUE={repr(cell.value)}  FORMULA={formula}  COLOR={color_label(color)}")

# ──────────────────────────────────────────────────────────────────────────────
# 2. PRODUCTION SCHEDULE SHEET — ALL cells with values/formulas
# ──────────────────────────────────────────────────────────────────────────────
print("\n\n=== PRODUCTION SCHEDULE SHEET ===\n")
prod_names = [n for n in wb_values.sheetnames if "production" in n.lower() or "prod" in n.lower()]
print(f"  Production-like sheets: {prod_names}")

for pname in prod_names:
    ws = wb_values[pname]
    wsf = wb_formulas[pname]
    print(f"\n--- Sheet: {pname} ---")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()
    
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            formula_cell = wsf[cell.coordinate]
            formula = formula_cell.value
            
            if val is None and (formula is None or not str(formula).startswith("=")):
                continue  # skip truly empty cells
            
            color = get_fill_color(wsf[cell.coordinate])
            color_str = color_label(color)
            
            # Print all non-empty cells
            if val is not None or (formula and str(formula).startswith("=")):
                print(f"  [{cell.coordinate}] VALUE={repr(val)!s:40s} FORMULA={repr(formula) if formula and str(formula).startswith('=') else '':60s} COLOR={color_str}")

# ──────────────────────────────────────────────────────────────────────────────
# 3. ASSUMPTIONS DASHBOARD — rows 1–30 and all RED cells
# ──────────────────────────────────────────────────────────────────────────────
print("\n\n=== ASSUMPTIONS DASHBOARD — HEADER ROWS + RED (HARD INPUT) CELLS ===\n")
dash_names = [n for n in wb_values.sheetnames if "assumption" in n.lower() or "dashboard" in n.lower()]
print(f"  Dashboard sheets: {dash_names}")

for dname in dash_names:
    ws = wb_values[dname]
    wsf = wb_formulas[dname]
    print(f"\n--- Sheet: {dname} (first 30 rows) ---")
    for row in ws.iter_rows(max_row=30):
        for cell in row:
            val = cell.value
            if val is not None and str(val).strip() != "":
                formula_cell = wsf[cell.coordinate]
                formula = formula_cell.value
                color = get_fill_color(wsf[cell.coordinate])
                print(f"  [{cell.coordinate}] VALUE={repr(val)!s:50s} FORMULA={repr(formula) if formula and str(formula).startswith('=') else '':60s} COLOR={color_label(color)}")
    
    print(f"\n--- Sheet: {dname} — ALL RED / DARK-RED (Hard Input) cells ---")
    red_colors = {"FFFF0000", "FFC00000", "FF8B0000", "FFFF4444"}
    for row in ws.iter_rows():
        for cell in row:
            formula_cell = wsf[cell.coordinate]
            color = get_fill_color(wsf[cell.coordinate])
            if color in red_colors:
                val = cell.value
                formula = formula_cell.value
                print(f"  [{cell.coordinate}] VALUE={repr(val)!s:50s} FORMULA={repr(formula) if formula and str(formula).startswith('=') else '':60s}")

# ──────────────────────────────────────────────────────────────────────────────
# 4. Specifically look for "Rehandling at CHP" and "Partings" in any sheet
# ──────────────────────────────────────────────────────────────────────────────
print("\n\n=== SEARCH: 'rehandl' and 'parting' across ALL sheets ===\n")
keywords = ["rehandl", "parting", "cHP rehandl", "% parting"]
for sheet_name in wb_values.sheetnames:
    ws = wb_values[sheet_name]
    wsf = wb_formulas[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""
            for kw in keywords:
                if kw.lower() in val.lower():
                    formula_cell = wsf[cell.coordinate]
                    formula = formula_cell.value
                    color = get_fill_color(wsf[cell.coordinate])
                    print(f"  Sheet={sheet_name!r:35s} [{cell.coordinate}] VALUE={repr(cell.value)!s:50s} FORMULA={repr(formula) if formula and str(formula).startswith('=') else '':60s} COLOR={color_label(color)}")

print("\n\nDone.")
