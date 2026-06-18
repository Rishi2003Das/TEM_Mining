import openpyxl

wb = openpyxl.load_workbook("/Users/rishidas/Documents/Internship_SRK/TEM_Mining/TEM_interface/IN1411_Radhikapur_Option 3_15 Mtpa_rev1_20260612.xlsx", data_only=True)

for name in ["Owner CAPEX", "MDO CAPEX", "Project CAPEX"]:
    if name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} ---")
        for r in range(1, 25):
            vals = [sheet.cell(r, c).value for c in range(1, 10)]
            if any(v is not None for v in vals):
                row_str = " | ".join([f"{c}:{str(v)}" for c, v in enumerate(vals, 1) if v is not None])
                print(f"Row {r:2d}: {row_str}")
